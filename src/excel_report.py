"""Update team tabs without modifying All Data or Summary."""

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.extract_tasks import extract_completed_tasks
from src.prepare_reporting_data import prepare_reporting_data
from src.team_mapping import TEAM_CONFIG


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "output"

PROTECTED_SHEETS = {
    "All Data",
    "Summary",
}

TEAM_HEADERS = [
    "Team",
    "Task Type",
    "Task ID",
    "Created Date",
    "Completed Date",
    "SLA Breached",
]

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


def find_existing_report() -> Path:
    """
    Find the existing KPI workbook in the output folder.

    The workflow expects exactly one KPI workbook so that it can
    update that workbook without creating or deleting another one.
    """

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    reports = sorted(
        report
        for report in OUTPUT_DIR.glob(
            "operations_kpi_report*.xlsx"
        )
        if not report.name.startswith("~$")
    )

    if not reports:
        raise FileNotFoundError(
            "No existing KPI workbook was found in the output folder. "
            "Place the completed workbook containing the All Data and "
            "Summary tabs in output before running the workflow."
        )

    if len(reports) > 1:
        report_names = "\n".join(
            f"- {report.name}"
            for report in reports
        )

        raise RuntimeError(
            "More than one KPI workbook was found in the output folder. "
            "Keep only the workbook that should be updated:\n"
            f"{report_names}"
        )

    return reports[0]


def snapshot_sheet(worksheet) -> dict:
    """
    Capture the contents of a protected worksheet.

    The snapshot is used to verify that the All Data and Summary
    worksheets remain unchanged after the workbook is saved.
    """

    values = tuple(
        tuple(
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
        )
        for row_number in range(
            1,
            worksheet.max_row + 1,
        )
    )

    merged_ranges = tuple(
        sorted(
            str(cell_range)
            for cell_range
            in worksheet.merged_cells.ranges
        )
    )

    return {
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "values": values,
        "merged_ranges": merged_ranges,
    }


def snapshot_protected_sheets(
    workbook,
) -> dict[str, dict]:
    """Capture the contents of All Data and Summary."""

    missing_sheets = (
        PROTECTED_SHEETS
        - set(workbook.sheetnames)
    )

    if missing_sheets:
        missing_names = ", ".join(
            sorted(missing_sheets)
        )

        raise KeyError(
            "The existing workbook is missing the following "
            f"protected sheet or sheets: {missing_names}"
        )

    return {
        sheet_name: snapshot_sheet(
            workbook[sheet_name]
        )
        for sheet_name in PROTECTED_SHEETS
    }


def create_table_name(
    team_name: str,
) -> str:
    """Create a valid Excel table name for a team."""

    cleaned_name = "".join(
        character
        for character in team_name
        if character.isalnum()
    )

    return f"{cleaned_name}Tasks"


def format_team_header(
    worksheet,
) -> None:
    """Format the team-tab header row."""

    for column_number, header in enumerate(
        TEAM_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 30

    column_widths = {
        "A": 18,
        "B": 34,
        "C": 16,
        "D": 21,
        "E": 21,
        "F": 16,
    }

    for column_letter, width in (
        column_widths.items()
    ):
        worksheet.column_dimensions[
            column_letter
        ].width = width


def remove_team_tables(
    worksheet,
) -> None:
    """
    Remove tables from the team data area.

    They are recreated after the new records are written. This does
    not affect tables or formulas on All Data or Summary.
    """

    for table_name in list(
        worksheet.tables.keys()
    ):
        table = worksheet.tables[
            table_name
        ]

        table_range = table.ref.upper()

        if table_range.startswith("A1:F"):
            del worksheet.tables[
                table_name
            ]


def clear_existing_team_data(
    worksheet,
) -> None:
    """
    Clear only columns A:F below the header.

    Columns G onward are not changed, so any manual KPI formulas or
    other content on the team worksheet are preserved.
    """

    if worksheet.max_row < 2:
        return

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=6,
    ):
        for cell in row:
            cell.value = None


def convert_excel_value(value):
    """Convert Pandas values into Excel-compatible values."""

    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def write_team_sheet(
    worksheet,
    team_df: pd.DataFrame,
    team_name: str,
) -> None:
    """
    Replace the task data in one team tab.

    Only columns A:F are updated.
    """

    remove_team_tables(
        worksheet
    )

    clear_existing_team_data(
        worksheet
    )

    format_team_header(
        worksheet
    )

    start_row = 2

    for row_number, row in enumerate(
        team_df.itertuples(index=False),
        start=start_row,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=convert_excel_value(
                row.team_name
            ),
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=convert_excel_value(
                row.task_type
            ),
        )

        worksheet.cell(
            row=row_number,
            column=3,
            value=convert_excel_value(
                row.task_id
            ),
        )

        worksheet.cell(
            row=row_number,
            column=4,
            value=convert_excel_value(
                row.created_at
            ),
        )

        worksheet.cell(
            row=row_number,
            column=5,
            value=convert_excel_value(
                row.completed_at
            ),
        )

        worksheet.cell(
            row=row_number,
            column=6,
            value=convert_excel_value(
                row.sla_breached
            ),
        )

    last_data_row = (
        len(team_df)
        + 1
    )

    if not team_df.empty:
        table_reference = (
            f"A1:F{last_data_row}"
        )

        table = Table(
            displayName=create_table_name(
                team_name
            ),
            ref=table_reference,
        )

        table.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        )

        worksheet.add_table(
            table
        )

    if last_data_row >= 2:
        for cell in worksheet[
            f"D2:D{last_data_row}"
        ][0]:
            cell.number_format = (
                "yyyy-mm-dd hh:mm"
            )

        for cell in worksheet[
            f"E2:E{last_data_row}"
        ][0]:
            cell.number_format = (
                "yyyy-mm-dd hh:mm"
            )

        for cell in worksheet[
            f"F2:F{last_data_row}"
        ][0]:
            cell.alignment = Alignment(
                horizontal="center",
            )


def verify_protected_sheets(
    original_snapshots: dict[str, dict],
    saved_workbook,
) -> None:
    """
    Confirm that All Data and Summary remain unchanged.

    The original workbook is not replaced if this check fails.
    """

    saved_snapshots = (
        snapshot_protected_sheets(
            saved_workbook
        )
    )

    changed_sheets = [
        sheet_name
        for sheet_name
        in PROTECTED_SHEETS
        if saved_snapshots[sheet_name]
        != original_snapshots[sheet_name]
    ]

    if changed_sheets:
        changed_names = ", ".join(
            sorted(changed_sheets)
        )

        raise RuntimeError(
            "The protected worksheet verification failed. "
            "The following worksheet contents changed during "
            f"the save process: {changed_names}"
        )


def export_kpi_report(
    reporting_df: pd.DataFrame,
) -> Path:
    """
    Update only the team tabs in the existing KPI workbook.

    All Data and Summary are loaded and preserved but are never
    written to by this workflow.
    """

    if reporting_df.empty:
        raise ValueError(
            "No reporting data is available to export."
        )

    report_path = find_existing_report()

    try:
        workbook = load_workbook(
            filename=report_path,
            data_only=False,
            keep_links=True,
        )

    except PermissionError as error:
        raise PermissionError(
            "Close the Excel workbook before running "
            "the workflow."
        ) from error

    protected_snapshots = (
        snapshot_protected_sheets(
            workbook
        )
    )

    for team_name in TEAM_CONFIG:
        if team_name not in workbook.sheetnames:
            worksheet = (
                workbook.create_sheet(
                    title=team_name
                )
            )
        else:
            worksheet = workbook[
                team_name
            ]

        team_df = (
            reporting_df[
                reporting_df["team_name"]
                == team_name
            ]
            .sort_values(
                by=[
                    "completed_at",
                    "task_type",
                    "task_id",
                ]
            )
            .reset_index(drop=True)
        )

        write_team_sheet(
            worksheet=worksheet,
            team_df=team_df,
            team_name=team_name,
        )

    # Request recalculation when the workbook is opened in Excel.
    # This does not write to the protected worksheets.
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    temporary_path = report_path.with_name(
        f"{report_path.stem}.temporary.xlsx"
    )

    try:
        workbook.save(
            temporary_path
        )

        saved_workbook = load_workbook(
            filename=temporary_path,
            data_only=False,
            keep_links=True,
        )

        verify_protected_sheets(
            original_snapshots=(
                protected_snapshots
            ),
            saved_workbook=(
                saved_workbook
            ),
        )

        saved_workbook.close()
        workbook.close()

        os.replace(
            temporary_path,
            report_path,
        )

    except PermissionError as error:
        if temporary_path.exists():
            temporary_path.unlink()

        raise PermissionError(
            "Close the Excel workbook before running "
            "the workflow."
        ) from error

    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()

        raise

    return report_path


if __name__ == "__main__":
    extracted_tasks_df = (
        extract_completed_tasks()
    )

    reporting_df = (
        prepare_reporting_data(
            extracted_tasks_df
        )
    )

    if reporting_df.empty:
        print(
            "No tasks were available "
            "for Excel reporting."
        )
    else:
        report_path = (
            export_kpi_report(
                reporting_df
            )
        )

        print(
            "Existing Excel report "
            "updated successfully:"
        )

        print(report_path)

        print(
            "The All Data and Summary tabs "
            "were preserved."
        )